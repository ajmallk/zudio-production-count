"""Report generation: PDF (ReportLab) and Excel (openpyxl)"""
import io
from datetime import datetime, timedelta
from django.utils import timezone
from django.db.models import Count, Q, Sum
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def get_date_range(period, date_from=None, date_to=None,
                   time_from=None, time_to=None,
                   sel_year=None, sel_month=None, sel_week=None, sel_day=None):
    """
    Returns (start_dt, end_dt) as timezone-aware datetimes.

    For 'custom':
      - If date_from / date_to are provided with optional time_from / time_to,
        the time components are applied directly.
      - time_from / time_to are strings in 'HH:MM' format.

    For 'year' / 'month' / 'week' / 'day':
      - sel_year  : int  — specific year  (e.g. 2025)
      - sel_month : str  — 'YYYY-MM'      (e.g. '2025-06')
      - sel_week  : str  — 'YYYY-Www'     (e.g. '2025-W23')
      - sel_day   : str  — 'YYYY-MM-DD'   (e.g. '2025-06-10')
    """
    import calendar
    now = timezone.now()
    today = now.date()

    if period == 'hour':
        return now - timedelta(hours=1), now

    elif period == 'day':
        if sel_day:
            try:
                chosen = datetime.strptime(sel_day, '%Y-%m-%d').date()
                start = timezone.make_aware(datetime.combine(chosen, datetime.min.time()))
                end   = timezone.make_aware(datetime.combine(chosen, datetime.max.time()))
                return start, end
            except ValueError:
                pass
        start = timezone.make_aware(datetime.combine(today, datetime.min.time()))
        return start, now

    elif period == 'week':
        if sel_week:
            try:
                # Parse ISO week string like '2025-W23'
                year_str, week_str = sel_week.split('-W')
                iso_year = int(year_str)
                iso_week = int(week_str)
                # Monday of that ISO week
                monday = datetime.fromisocalendar(iso_year, iso_week, 1).date()
                sunday = monday + timedelta(days=6)
                start = timezone.make_aware(datetime.combine(monday, datetime.min.time()))
                end   = timezone.make_aware(datetime.combine(sunday, datetime.max.time()))
                return start, end
            except (ValueError, AttributeError):
                pass
        start_date = today - timedelta(days=today.weekday())
        return timezone.make_aware(datetime.combine(start_date, datetime.min.time())), now

    elif period == 'month':
        if sel_month:
            try:
                parsed = datetime.strptime(sel_month, '%Y-%m')
                y, m = parsed.year, parsed.month
                last_day = calendar.monthrange(y, m)[1]
                start = timezone.make_aware(datetime(y, m, 1, 0, 0, 0))
                end   = timezone.make_aware(datetime(y, m, last_day, 23, 59, 59))
                return start, end
            except ValueError:
                pass
        start = today.replace(day=1)
        return timezone.make_aware(datetime.combine(start, datetime.min.time())), now

    elif period == 'year':
        if sel_year:
            try:
                y = int(sel_year)
                start = timezone.make_aware(datetime(y, 1, 1, 0, 0, 0))
                end   = timezone.make_aware(datetime(y, 12, 31, 23, 59, 59))
                return start, end
            except ValueError:
                pass
        start = today.replace(month=1, day=1)
        return timezone.make_aware(datetime.combine(start, datetime.min.time())), now

    elif period == 'custom' and date_from and date_to:
        # Parse optional times
        try:
            t_from = datetime.strptime(time_from, '%H:%M').time() if time_from else datetime.min.time()
        except ValueError:
            t_from = datetime.min.time()
        try:
            t_to = datetime.strptime(time_to, '%H:%M').time() if time_to else datetime.max.time()
        except ValueError:
            t_to = datetime.max.time()
        return (
            timezone.make_aware(datetime.combine(date_from, t_from)),
            timezone.make_aware(datetime.combine(date_to, t_to)),
        )
    else:
        start = today.replace(day=1)
        return timezone.make_aware(datetime.combine(start, datetime.min.time())), now


def build_report_data(period='month', date_from=None, date_to=None,
                      time_from=None, time_to=None,
                      sel_year=None, sel_month=None, sel_week=None, sel_day=None):
    from .models import QRCode, ScanEvent

    start_dt, end_dt = get_date_range(
        period, date_from, date_to, time_from, time_to,
        sel_year=sel_year, sel_month=sel_month,
        sel_week=sel_week, sel_day=sel_day,
    )

    # Production report: QR codes printed by date/article/size (within range)
    qr_qs = QRCode.objects.filter(created_at__range=(start_dt, end_dt))
    production_data = []
    for qr in qr_qs.values('article_number', 'size', 'created_at__date').annotate(
        count=Count('id')
    ).order_by('created_at__date', 'article_number', 'size'):
        production_data.append({
            'date':    str(qr['created_at__date']),
            'article': qr['article_number'],
            'size':    qr['size'],
            'printed': qr['count'],
        })

    # Total records and total production count for the period
    total_records         = len(production_data)
    total_production_count = qr_qs.count()

    # Stock report: scanned vs unscanned by article/size (all-time)
    stock_data = []
    for item in QRCode.objects.values('article_number', 'size').distinct().order_by('article_number', 'size'):
        qs = QRCode.objects.filter(article_number=item['article_number'], size=item['size'])
        total   = qs.count()
        scanned = qs.filter(is_scanned=True).count()
        pct     = round((scanned / total * 100), 1) if total > 0 else 0
        stock_data.append({
            'article':    item['article_number'],
            'size':       item['size'],
            'printed':    total,
            'scanned':    scanned,
            'unscanned':  total - scanned,
            'completion': pct,
        })

    # Audit log (within range)
    audit_data = []
    for event in (ScanEvent.objects
                  .filter(scanned_at__range=(start_dt, end_dt))
                  .select_related('qr_code')
                  .order_by('-scanned_at')[:500]):
        audit_data.append({
            'timestamp': event.scanned_at.strftime('%Y-%m-%d %H:%M:%S'),
            'qr_id':    str(event.qr_code.qr_id) if event.qr_code else 'N/A',
            'size':     event.qr_code.size if event.qr_code else 'N/A',
            'article':  event.qr_code.article_number if event.qr_code else 'N/A',
            'status':   event.get_status_display(),
            'device':   event.device_info or 'Unknown',
        })

    return (
        production_data, stock_data, audit_data,
        start_dt, end_dt,
        total_records, total_production_count,
    )


def generate_excel_report(period='month', date_from=None, date_to=None,
                           time_from=None, time_to=None,
                           sel_year=None, sel_month=None, sel_week=None, sel_day=None):
    (production_data, stock_data, audit_data,
     start_dt, end_dt,
     total_records, total_production_count) = build_report_data(
        period, date_from, date_to, time_from, time_to,
        sel_year=sel_year, sel_month=sel_month,
        sel_week=sel_week, sel_day=sel_day,
    )

    wb = openpyxl.Workbook()

    # ── Styles ────────────────────────────────────────────────────────────
    header_font   = Font(bold=True, color="FFFFFF", size=12)
    header_fill   = PatternFill("solid", fgColor="1a1a2e")
    meta_font     = Font(bold=True, size=11)
    meta_fill     = PatternFill("solid", fgColor="E8EAFF")
    center_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align    = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    border        = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    def write_meta_row(ws, row_num, label, value):
        """Write a highlighted metadata row."""
        ws.cell(row=row_num, column=1, value=label).font  = meta_font
        ws.cell(row=row_num, column=1).fill               = meta_fill
        ws.cell(row=row_num, column=1).alignment          = left_align
        ws.cell(row=row_num, column=2, value=value).font  = meta_font
        ws.cell(row=row_num, column=2).alignment          = left_align
        ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=4)

    def style_sheet(ws, headers, data_rows, col_widths, start_row=1):
        ws.append(headers)
        for cell in ws[start_row]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center_align
            cell.border    = border
        for row_data in data_rows:
            ws.append(row_data)
        for row in ws.iter_rows(min_row=start_row + 1):
            for cell in row:
                cell.alignment = center_align
                cell.border    = border
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[start_row].height = 25

    # ── Range description ─────────────────────────────────────────────────
    range_label = f"{start_dt.strftime('%d %b %Y %H:%M')} — {end_dt.strftime('%d %b %Y %H:%M')}"

    # ── Production Sheet ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Production Report"
    # Meta rows
    write_meta_row(ws1, 1, "Report Period", range_label)
    write_meta_row(ws1, 2, "Total Record Groups", str(total_records))
    write_meta_row(ws1, 3, "Total QR Codes Produced", str(total_production_count))
    ws1.append([])  # blank spacer row
    style_sheet(
        ws1,
        ['Date', 'Article Number', 'Size', 'QR Codes Printed'],
        [[d['date'], d['article'], d['size'], d['printed']] for d in production_data],
        [15, 20, 10, 18],
        start_row=5,
    )

    # ── Stock Sheet ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Stock Report")
    write_meta_row(ws2, 1, "Report Period", range_label)
    write_meta_row(ws2, 2, "Total QR Codes Produced (period)", str(total_production_count))
    ws2.append([])
    style_sheet(
        ws2,
        ['Article Number', 'Size', 'Printed', 'Scanned', 'Unscanned', 'Completion %'],
        [[d['article'], d['size'], d['printed'], d['scanned'], d['unscanned'], d['completion']] for d in stock_data],
        [20, 10, 12, 12, 12, 15],
        start_row=4,
    )

    # ── Audit Sheet ───────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Audit Log")
    write_meta_row(ws3, 1, "Report Period", range_label)
    ws3.append([])
    style_sheet(
        ws3,
        ['Timestamp', 'QR ID', 'Size', 'Article', 'Status', 'Device'],
        [[d['timestamp'], d['qr_id'], d['size'], d['article'], d['status'], d['device']] for d in audit_data],
        [20, 38, 10, 20, 16, 30],
        start_row=3,
    )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_pdf_report(period='month', date_from=None, date_to=None,
                        time_from=None, time_to=None,
                        sel_year=None, sel_month=None, sel_week=None, sel_day=None):
    (production_data, stock_data, audit_data,
     start_dt, end_dt,
     total_records, total_production_count) = build_report_data(
        period, date_from, date_to, time_from, time_to,
        sel_year=sel_year, sel_month=sel_month,
        sel_week=sel_week, sel_day=sel_day,
    )

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4),
                            rightMargin=1*cm, leftMargin=1*cm,
                            topMargin=1.5*cm, bottomMargin=1*cm)

    styles     = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', fontSize=16, fontName='Helvetica-Bold',
                                  textColor=colors.HexColor('#1a1a2e'),
                                  alignment=TA_CENTER, spaceAfter=4)
    sub_style   = ParagraphStyle('Sub', fontSize=10, fontName='Helvetica',
                                  textColor=colors.grey,
                                  alignment=TA_CENTER, spaceAfter=4)
    meta_style  = ParagraphStyle('Meta', fontSize=10, fontName='Helvetica-Bold',
                                  textColor=colors.HexColor('#1a1a2e'),
                                  alignment=TA_CENTER, spaceAfter=8)
    section_style = ParagraphStyle('Section', fontSize=13, fontName='Helvetica-Bold',
                                    textColor=colors.HexColor('#e94560'),
                                    spaceBefore=12, spaceAfter=6)

    header_bg = colors.HexColor('#1a1a2e')
    alt_row   = colors.HexColor('#f0f4ff')
    meta_bg   = colors.HexColor('#E8EAFF')

    def make_table(headers, rows):
        data = [headers] + rows
        t    = Table(data, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), header_bg),
            ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 0), (-1, 0), 9),
            ('ALIGN',      (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTSIZE',   (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, alt_row]),
            ('GRID',       (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
            ('ROWHEIGHT',  (0, 0), (-1, -1), 18),
        ]))
        return t

    def make_summary_table(rows):
        """Highlighted summary / metadata table."""
        t = Table(rows, colWidths=[9*cm, 8*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), meta_bg),
            ('FONTNAME',   (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME',   (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE',   (0, 0), (-1, -1), 9),
            ('ALIGN',      (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID',       (0, 0), (-1, -1), 0.5, colors.HexColor('#c7d2fe')),
            ('ROWHEIGHT',  (0, 0), (-1, -1), 18),
        ]))
        return t

    range_label = f"{start_dt.strftime('%d %b %Y  %H:%M')}  —  {end_dt.strftime('%d %b %Y  %H:%M')}"

    story = []
    story.append(Paragraph("VKC Footwear — Production & QR Stock Report", title_style))
    story.append(Paragraph(f"Generated: {timezone.now().strftime('%d %b %Y %H:%M:%S')}", sub_style))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e94560')))
    story.append(Spacer(1, 6))

    # ── Summary metadata block ────────────────────────────────────────────
    summary_rows = [
        ["Report Period",            range_label],
        ["Total Record Groups",      str(total_records)],
        ["Total QR Codes Produced",  str(total_production_count)],
    ]
    story.append(make_summary_table(summary_rows))
    story.append(Spacer(1, 12))

    # ── Production ────────────────────────────────────────────────────────
    story.append(Paragraph("Production Report (QR Codes Printed)", section_style))
    if production_data:
        rows = [[d['date'], d['article'], str(d['size']), str(d['printed'])] for d in production_data]
        story.append(make_table(['Date', 'Article Number', 'Size', 'QR Printed'], rows))
    else:
        story.append(Paragraph("No production data for this period.", styles['Normal']))

    story.append(Spacer(1, 16))

    # ── Stock ─────────────────────────────────────────────────────────────
    story.append(Paragraph("Stock Report (Scanned vs Unscanned)", section_style))
    if stock_data:
        rows = [[d['article'], str(d['size']), str(d['printed']), str(d['scanned']),
                 str(d['unscanned']), f"{d['completion']}%"] for d in stock_data]
        story.append(make_table(['Article Number', 'Size', 'Printed', 'Scanned', 'Unscanned', 'Completion %'], rows))
    else:
        story.append(Paragraph("No stock data available.", styles['Normal']))

    story.append(Spacer(1, 16))

    # ── Audit ─────────────────────────────────────────────────────────────
    story.append(Paragraph("Audit Log (Scan Events)", section_style))
    if audit_data:
        rows = [[d['timestamp'], d['qr_id'][:16] + '...', str(d['size']),
                 d['article'], d['status'], d['device'][:25]] for d in audit_data]
        story.append(make_table(['Timestamp', 'QR ID', 'Size', 'Article', 'Status', 'Device'], rows))
    else:
        story.append(Paragraph("No scan events in this period.", styles['Normal']))

    doc.build(story)
    output.seek(0)
    return output


def generate_batch_pdf(batch, width_mm=None, height_mm=None):
    """Generate printable PDF sheet with all QR codes for a batch"""
    from .models import QRCode
    from reportlab.platypus import Image as RLImage, PageBreak
    from reportlab.lib.units import mm

    qr_codes = QRCode.objects.filter(batch=batch).order_by('size')
    output   = io.BytesIO()
    
    is_thermal = width_mm and height_mm
    
    if is_thermal:
        page_size = (width_mm * mm, height_mm * mm)
        doc = SimpleDocTemplate(output, pagesize=page_size,
                                rightMargin=0, leftMargin=0,
                                topMargin=0, bottomMargin=0)
    else:
        doc = SimpleDocTemplate(output, pagesize=A4,
                                rightMargin=1*cm, leftMargin=1*cm,
                                topMargin=1.5*cm, bottomMargin=1*cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('T', fontSize=14, fontName='Helvetica-Bold',
                                  textColor=colors.HexColor('#1a1a2e'),
                                  alignment=TA_CENTER, spaceAfter=4)
    info_style  = ParagraphStyle('I', fontSize=8, fontName='Helvetica',
                                  textColor=colors.grey,
                                  alignment=TA_CENTER, spaceAfter=12)

    story = []
    
    if not is_thermal:
        story.append(Paragraph("VKC Footwear — QR Code Sheet", title_style))
        story.append(Paragraph(
            f"Batch: {batch.batch_id} | Generated: {batch.created_at.strftime('%d %b %Y %H:%M')}",
            info_style
        ))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e94560')))
        story.append(Spacer(1, 8))

    name_style_pdf = ParagraphStyle(
        'nl', fontSize=10, fontName='Helvetica-Bold',
        textColor=colors.HexColor('#4f46e5'), alignment=TA_CENTER
    )

    if is_thermal:
        # Thermal mode: 1 QR per page, exact size
        # Calculate max QR size that fits while leaving some margin
        margin = 2 * mm
        avail_w = (width_mm * mm) - margin*2
        avail_h = (height_mm * mm) - margin*2
        # Deduct some space for text in thermal mode
        text_space = 8 * mm # approx space for text
        QR_SIZE = min(avail_w, avail_h - text_space)
        if QR_SIZE < 10*mm:
             QR_SIZE = 10*mm
             
        for qr in qr_codes:
            if qr.qr_image and qr.qr_image.path and __import__('os').path.exists(qr.qr_image.path):
                img = RLImage(qr.qr_image.path, width=QR_SIZE, height=QR_SIZE)
                # Build a single column table for the page
                cell_content = [
                    Spacer(1, margin),
                    img
                ]
                story.append(Table([[c] for c in cell_content], colWidths=[width_mm * mm], style=[
                    ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ]))
                story.append(PageBreak())
            else:
                story.append(Paragraph(f"QR Size {qr.size}", ParagraphStyle('c', alignment=TA_CENTER)))
                story.append(PageBreak())
    else:
        # Build grid: 4 per row
        QR_SIZE      = 4.5 * cm
        items_per_row = 4
        row_data      = []
        current_row   = []

        for qr in qr_codes:
            if qr.qr_image and qr.qr_image.path and __import__('os').path.exists(qr.qr_image.path):
                cell_content = [
                    RLImage(qr.qr_image.path, width=QR_SIZE, height=QR_SIZE),
                    Paragraph(f"<b>Size {qr.size}</b>",
                              ParagraphStyle('c', fontSize=8, alignment=TA_CENTER)),
                    Paragraph(qr.article_number,
                              ParagraphStyle('c2', fontSize=6, textColor=colors.grey, alignment=TA_CENTER)),
                ]
                # Name — large and prominent in the print layout
                if qr.name:
                    cell_content.append(
                        Paragraph(qr.name,
                                  name_style_pdf)
                    )
            else:
                cell_content = [Paragraph(f"QR\nSize {qr.size}",
                                           ParagraphStyle('c', fontSize=8, alignment=TA_CENTER))]

            current_row.append(cell_content)
            if len(current_row) == items_per_row:
                row_data.append(current_row)
                current_row = []

        if current_row:
            while len(current_row) < items_per_row:
                current_row.append([Spacer(1, 1)])
            row_data.append(current_row)

        if row_data:
            col_w = [QR_SIZE + 0.5*cm] * items_per_row
            row_h = [QR_SIZE + 2.0*cm] * len(row_data)   # extra height for name row
            t = Table(row_data, colWidths=col_w, rowHeights=row_h)
            t.setStyle(TableStyle([
                ('ALIGN',  (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID',   (0, 0), (-1, -1), 0.5, colors.HexColor('#dddddd')),
                ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ]))
            story.append(t)

    doc.build(story)
    output.seek(0)
    return output
